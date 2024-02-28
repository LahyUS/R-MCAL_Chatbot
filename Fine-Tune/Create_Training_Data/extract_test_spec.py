import pandas as pd
import os
import json
import re
from openpyxl import load_workbook

def find_section_column(df, section_marker_4th_row, section_marker_2nd_row):
    # Check 4th row for section marker
    section_column_4th_row = df.loc[4][df.loc[4] == section_marker_4th_row]
    if not section_column_4th_row.empty:
        section_column = section_column_4th_row.index[0]
    else:
        # Check 2nd row for section marker as fallback
        section_column_2nd_row = df.loc[2][df.loc[2] == section_marker_2nd_row]
        if not section_column_2nd_row.empty:
            section_column = section_column_2nd_row.index[0]
        else:
            section_column = None

    return section_column


def extract_function_name(df):
    # Extract function name from cell A1 if it matches the specified format
    function_name = None
    if df.iloc[0, 0]:
        # Check for "Tested function : <function_name>" pattern
        #function_name_match   = re.match(r"Tested function\s*:\s*(\S+)", df.iloc[0, 0])
        tested_function_match = re.match(r"Tested function\s*:\s*(\S+)", df.iloc[0, 0])
        if tested_function_match:
            function_name = tested_function_match.group(1)
        else:
            # Check for ": <function_name>" pattern
            colon_function_match = re.match(r":\s*(\S+)", df.iloc[0, 0])
            if colon_function_match:
                function_name = colon_function_match.group(1)
    return function_name


def extract_input_output(excel_file):
    # Read Excel file into pandas dataframe
    xls = pd.ExcelFile(excel_file)
    sheet_names = xls.sheet_names

    data = {}
    for sheet_name in sheet_names:
        test_sheet = {}
        # Check if sheet name is a test sheet (numeric)
        # Check if sheet name matches the pattern of digits followed by characters
        if sheet_name.isdigit() or re.match(r'^\d+[a-zA-Z]+$', sheet_name):
            df = pd.read_excel(excel_file, sheet_name=sheet_name, header=None)

            # Extract function name from cell A1
            function_name = extract_function_name(df)

            # Find columns containing Input and Output sections
            input_column = find_section_column(df, 'Input', 'Target test input')
            output_column = find_section_column(df, 'Output(Expected value)', 'Target output value')

            if input_column is not None and output_column is not None:
                ### VERSION 1
                # # Extract Input and Output columns excluding rows 1 to 5
                # input_data = df.loc[5:, input_column:output_column-1]
                # output_data = df.loc[5:, output_column:]

                # data[sheet_name] = {
                #     'Input': input_data.to_dict(orient='list'),
                #     'Output': output_data.to_dict(orient='list')
                # }

                #### VERSION 2
                # input_data = df.loc[5:, input_column:output_column-1].fillna('').astype(str)
                # output_data = df.loc[5:, output_column:].fillna('').astype(str)

                # input_dict = {str(idx): row.tolist() for idx, row in input_data.iterrows()}
                # output_dict = {str(idx): row.tolist() for idx, row in output_data.iterrows()}

                # test_data = {}
                # for key in input_dict.keys():
                #     test_data[key] = {
                #         'Input': {key: input_dict[key]},
                #         'Output': {key: output_dict[key]}
                #     }

                type_row = df.loc[6, input_column:].fillna('').astype(str).tolist()
                range_row = df.loc[8, input_column:].fillna('').astype(str).tolist()
                name_row = df.loc[9, input_column:].fillna('').astype(str).tolist()

                input_data = df.loc[11:, input_column:output_column-1].fillna('').astype(str)
                output_data = df.loc[11:, output_column:].fillna('').astype(str)

                for row_idx, (row_key, row) in enumerate(input_data.iterrows(), start=1):
                    input_values = row.tolist()
                    output_values = output_data.iloc[row_idx-1].tolist()
                    input_dict = {}
                    output_dict = {}
                    iter = 0
                    for input_idx, val in enumerate(input_values):
                        input_params = {}
                        input_params['Type']  = type_row[iter]
                        input_params['Range'] = range_row[iter]
                        input_params['Name']  = name_row[iter]
                        input_params['Value']  = val
                        iter += 1
                        input_dict[f"Input_Param_{input_idx:03d}"] = input_params

                    for output_idx, val in enumerate(output_values):
                        output_params = {}
                        output_params['Type']  = type_row[iter]
                        output_params['Range'] = range_row[iter]
                        output_params['Name']  = name_row[iter]
                        output_params['Value']  = val
                        iter += 1
                        output_dict[f"Output_Param_{output_idx:03d}"] = output_params
                        
                    test_data = {
                        'Input': input_dict,
                        'Output': output_dict
                    }

                    test_sheet[f"Testcase_{str(row_idx).zfill(3)}"] = test_data  # Pad sheet_name with leading zeros
                
                #data[f"{function_name}_{sheet_name.zfill(3)}"] = test_sheet  # Pad sheet_name with leading zeros
                data[f"{function_name}"] = test_sheet  # Pad sheet_name with leading zeros
    return data


def split_horizontal_merged_cells(sheet, start_row, end_row):
    merged_ranges_copy = sheet.merged_cells.ranges.copy()

    for merged_range in merged_ranges_copy:
        min_row, min_col, max_row, max_col = merged_range.min_row, merged_range.min_col, merged_range.max_row, merged_range.max_col
        if min_row >= start_row and max_row <= end_row:
            merged_value = sheet.cell(row=min_row, column=min_col).value

            sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row=row, column=col, value=merged_value)

def process_excel_file(file_path, modified_path, start_row, end_row):
    try:
        wb = load_workbook(file_path)

        for sheet_name in wb.sheetnames:
            # Check if sheet name matches the pattern of digits followed by characters
            if sheet_name.isdigit() or re.match(r'^\d+[a-zA-Z]+$', sheet_name):
                sheet = wb[sheet_name]
                split_horizontal_merged_cells(sheet, start_row, end_row)

        raw_filename = os.path.basename(file_path)
        filename = os.path.splitext(raw_filename)[0]
        output_path = os.path.join(modified_path, f"{filename}.xlsx")
        wb.save(output_path)
        
        print(f"File '{file_path}' processed. Modified file saved as '{output_path}'")

    except Exception as e:
        print(f"An error occurred: {str(e)}")

PRE_PROCESS_TEST_SPEC = False
PROCESS_TEST_SPEC = True

if __name__ == "__main__":
    input_folder = 'C:/Users/rvc_sw_mss1_common/Desktop/Create_Training_Data/DUT_Test_Spec'  # Replace with your folder path containing Excel files
    modified_input_folder = 'C:/Users/rvc_sw_mss1_common/Desktop/Create_Training_Data/DUT_Test_Spec/Modified'  # Replace with your folder path containing Excel files
    output_folder = 'C:/Users/rvc_sw_mss1_common/Desktop/Create_Training_Data/JSON_Test_Spec'  # Replace with desired JSON output file
    excel_file_path = 'C:/Users/rvc_sw_mss1_common/Desktop/Create_Training_Data/DUT_Test_Spec/RH850_X2x_ADC_DUT_TS.xlsm'
    modified_folder = 'C:/Users/rvc_sw_mss1_common/Desktop/Create_Training_Data/DUT_Test_Spec/Modified'

    all_files = os.listdir(input_folder)
    excel_files = [file for file in all_files if file.endswith('.xlsm') or file.endswith('.xlsx') or file.endswith('.xls')]

    if PRE_PROCESS_TEST_SPEC:
        for file in excel_files:
            file_path = os.path.join(input_folder, file)
            # Read Excel file into pandas dataframe
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names

            data = {}
            process_excel_file(file_path, modified_folder, 6, 10)

    elif PROCESS_TEST_SPEC:   
        modifed_files = os.listdir(modified_input_folder)
        modifed_excel_files = [file for file in modifed_files if file.endswith('.xlsm') or file.endswith('.xlsx') or file.endswith('.xls')]
        for file in modifed_excel_files:
            file_path = os.path.join(modified_input_folder, file)
            extracted_data = extract_input_output(file_path)

            file_folder = os.path.splitext(file)[0]
            file_output_folder = os.path.join(output_folder, file_folder)

            if not os.path.exists(file_output_folder):
                os.makedirs(file_output_folder)

            for sheet_name, data in extracted_data.items():
                sheet_json_file = os.path.join(file_output_folder, f"{sheet_name}.json")

                save_data = {}
                for test_case_key, test_case_data in data.items():
                    # Replace NaN values with empty strings
                    save_data['Input'] = {key: [val if isinstance(val, str) else '' for val in values] for key, values in test_case_data['Input'].items()}
                    save_data['Output'] = {key: [val if isinstance(val, str) else '' for val in values] for key, values in test_case_data['Output'].items()}

                with open(sheet_json_file, 'w') as json_file:
                    json.dump(data, json_file, indent=4)
